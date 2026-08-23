"""
Student module tests — export isolation and student CRUD isolation.
"""
import io
import openpyxl
from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse

from departments.models import Department
from accounts.models import StaffProfile
from .models import Student


class StudentExportTestCase(TestCase):
    def setUp(self):
        self.dept_cse = Department.objects.create(code='CSE', name='Computer Science')
        self.dept_ece = Department.objects.create(code='ECE', name='Electronics')

        self.hod_cse = StaffProfile.objects.create(
            user=User.objects.create_user(username='exp_hod', password='password123'),
            employee_id='E-HOD', department=self.dept_cse, role=StaffProfile.ROLE_HOD,
        )
        self.teacher_cse = StaffProfile.objects.create(
            user=User.objects.create_user(username='exp_teacher', password='password123'),
            employee_id='E-TCH', department=self.dept_cse, role=StaffProfile.ROLE_TEACHER,
        )
        self.super_admin = StaffProfile.objects.create(
            user=User.objects.create_user(username='exp_admin', password='password123', is_staff=True, is_superuser=True),
            employee_id='E-ADM', role=StaffProfile.ROLE_SUPER_ADMIN,
        )

        for code, dept, reg in [('CS-001', self.dept_cse, 'CS-001'), ('EC-001', self.dept_ece, 'EC-001')]:
            Student.objects.create(
                register_number=reg, name=f'Student {code}', department=dept,
                batch='2023-2027', academic_year='2026-27', year=4, semester=7, section='A',
            )

    def test_teacher_cannot_export(self):
        client = Client()
        client.login(username='exp_teacher', password='password123')
        response = client.get(reverse('students:export'))
        self.assertEqual(response.status_code, 403)

    def test_hod_export_contains_only_own_department(self):
        client = Client()
        client.login(username='exp_hod', password='password123')
        response = client.get(reverse('students:export'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        regs = [row[0] for row in rows[1:]]
        self.assertIn('CS-001', regs)
        self.assertNotIn('EC-001', regs)

    def test_hod_cannot_force_other_department_export(self):
        """Even if a HOD passes another department id in the URL, export is scoped."""
        client = Client()
        client.login(username='exp_hod', password='password123')
        response = client.get(reverse('students:export'), {'department': self.dept_ece.id})
        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        regs = [row[0] for row in rows[1:]]
        self.assertIn('CS-001', regs)
        self.assertNotIn('EC-001', regs)

    def test_super_admin_can_export_all(self):
        client = Client()
        client.login(username='exp_admin', password='password123')
        response = client.get(reverse('students:export'))
        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        regs = [row[0] for row in rows[1:]]
        self.assertIn('CS-001', regs)
        self.assertIn('EC-001', regs)

    def test_export_with_filters(self):
        Student.objects.create(
            register_number='CS-002', name='Second Batch', department=self.dept_cse,
            batch='2022-2026', academic_year='2026-27', year=3, semester=5, section='B',
        )
        client = Client()
        client.login(username='exp_hod', password='password123')
        response = client.get(reverse('students:export'), {'batch': '2023-2027'})
        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        regs = [row[0] for row in list(ws.iter_rows(values_only=True))[1:]]
        self.assertIn('CS-001', regs)
        self.assertNotIn('CS-002', regs)


class StudentCRUDIsolationTestCase(TestCase):
    def setUp(self):
        self.dept_cse = Department.objects.create(code='CSE', name='Computer Science')
        self.dept_ece = Department.objects.create(code='ECE', name='Electronics')
        self.hod_ece = StaffProfile.objects.create(
            user=User.objects.create_user(username='crud_ece', password='password123'),
            employee_id='C-ECE', department=self.dept_ece, role=StaffProfile.ROLE_HOD,
        )

    def test_hod_creates_student_in_own_department_only(self):
        client = Client()
        client.login(username='crud_ece', password='password123')
        response = client.post(reverse('students:add'), {
            'register_number': 'HACK-001',
            'name': 'Hacker',
            'department': self.dept_cse.id,
            'batch': '2023-2027',
            'academic_year': '2026-27',
            'year': 4,
            'semester': 7,
            'section': 'A',
            'status': 'active',
        })
        self.assertNotEqual(response.status_code, 200)
        student = Student.objects.get(register_number='HACK-001')
        self.assertEqual(student.department, self.dept_ece)
