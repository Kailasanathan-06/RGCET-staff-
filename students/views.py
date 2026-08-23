from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from accounts.mixins import (
    DepartmentAccessMixin, RoleRequiredMixin, HODOrAboveMixin, LoginAndActiveRequiredMixin
)
from accounts.models import StaffProfile
from audit.models import AuditLog
from .models import Student
from django import forms


class StudentForm(forms.ModelForm):
    class Meta:
        model = Student
        fields = [
            'register_number', 'name', 'email', 'phone', 
            'batch', 'academic_year', 'year', 'semester', 'section', 'status'
        ]
        widgets = {
            'register_number': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. 21CS001'}),
            'name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. Anil Kumar'}),
            'email': forms.EmailInput(attrs={'class': 'form-control', 'placeholder': 'e.g. anil@rgcet.edu'}),
            'phone': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. 9876543210'}),
            'batch': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. 2023-2027'}),
            'academic_year': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. 2026-27'}),
            'year': forms.Select(attrs={'class': 'form-select'}),
            'semester': forms.Select(attrs={'class': 'form-select'}),
            'section': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. A'}),
            'status': forms.Select(attrs={'class': 'form-select'}),
        }


class StudentListView(DepartmentAccessMixin, ListView):
    """
    Lists students in the user's department.
    Enforces strict isolation.
    """
    model = Student
    template_name = 'students/student_list.html'
    context_object_name = 'students'
    paginate_by = 25

    def get_queryset(self):
        qs = super().get_queryset().select_related('department')
        
        # Search
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(register_number__icontains=q) |
                Q(email__icontains=q)
            )

        # Filters
        batch = self.request.GET.get('batch')
        if batch:
            qs = qs.filter(batch=batch)

        year = self.request.GET.get('year')
        if year:
            qs = qs.filter(year=year)

        section = self.request.GET.get('section')
        if section:
            qs = qs.filter(section=section)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Distinct values for filters scoped to department
        dept = self.get_user_department()
        base_qs = Student.objects.all()
        if dept:
            base_qs = base_qs.filter(department=dept)
            
        ctx['batches'] = base_qs.values_list('batch', flat=True).distinct()
        ctx['sections'] = base_qs.values_list('section', flat=True).distinct()
        ctx['current_filters'] = self.request.GET
        ctx['export_querystring'] = self.request.GET.urlencode() if self.request.GET else ''
        return ctx


class StudentDetailView(DepartmentAccessMixin, DetailView):
    model = Student
    template_name = 'students/student_detail.html'
    context_object_name = 'student'


class StudentCreateView(HODOrAboveMixin, DepartmentAccessMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'students/student_form.html'
    success_url = reverse_lazy('students:list')

    def form_valid(self, form):
        student = form.save(commit=False)
        profile = self.request.user.profile
        # Scope student to the user's department (HOD's department)
        if not profile.is_super_admin:
            student.department = profile.department
        
        student.save()
        AuditLog.log(
            user=self.request.user,
            action=AuditLog.ACTION_UPLOAD,
            description=f"Created student record: {student.register_number}",
            obj=student,
            ip_address=self.request.client_ip
        )
        messages.success(self.request, f"✓ Student '{student.name}' added successfully.")
        return redirect(self.success_url)


class StudentUpdateView(HODOrAboveMixin, DepartmentAccessMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'students/student_form.html'
    success_url = reverse_lazy('students:list')

    def form_valid(self, form):
        response = super().form_valid(form)
        AuditLog.log(
            user=self.request.user,
            action=AuditLog.ACTION_UPDATE,
            description=f"Updated student record: {self.object.register_number}",
            obj=self.object,
            ip_address=self.request.client_ip
        )
        messages.success(self.request, "✓ Student record updated successfully.")
        return response


class StudentDeleteView(HODOrAboveMixin, DepartmentAccessMixin, DeleteView):
    model = Student
    template_name = 'students/student_confirm_delete.html'
    success_url = reverse_lazy('students:list')

    def form_valid(self, form):
        student = self.get_object()
        AuditLog.log(
            user=self.request.user,
            action=AuditLog.ACTION_DELETE,
            description=f"Deleted student record: {student.register_number}",
            obj=student,
            ip_address=self.request.client_ip
        )
        messages.success(self.request, f"Student record '{student.register_number}' has been deleted.")
        return super().form_valid(form)


def export_students(request):
    """
    Export students to an .xlsx file.

    Access: HOD (own department) and Super Admin (any department).
    Filters: department, batch, year, semester, section, academic_year.
    The department filter is locked to the HOD's own department — even
    if a HOD crafts a URL with another department id, it is overridden.
    """
    profile = request.user.profile

    if not profile or profile.role not in (StaffProfile.ROLE_SUPER_ADMIN, StaffProfile.ROLE_HOD):
        raise PermissionDenied("You are not authorized to export student data.")

    qs = Student.objects.select_related('department')

    # Department isolation: HOD is forced to their own department.
    dept_id = request.GET.get('department')
    if profile.is_super_admin and dept_id:
        qs = qs.filter(department_id=dept_id)
    elif not profile.is_super_admin:
        qs = qs.filter(department=profile.department)

    filters = {}
    for param, field in [
        ('batch', 'batch'),
        ('year', 'year'),
        ('semester', 'semester'),
        ('section', 'section'),
        ('academic_year', 'academic_year'),
    ]:
        value = request.GET.get(param)
        if value:
            filters[field] = value

    qs = qs.filter(**filters).order_by('department__code', 'batch', 'year', 'section', 'name')

    # ── Build Excel workbook ───────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    headers = [
        'Register Number', 'Name', 'Department', 'Batch', 'Academic Year',
        'Year', 'Semester', 'Section', 'Email', 'Phone', 'Status'
    ]
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, student in enumerate(qs, start=2):
        values = [
            student.register_number,
            student.name,
            student.department.code,
            student.batch,
            student.academic_year,
            student.year,
            student.semester,
            student.section,
            student.email,
            student.phone,
            student.get_status_display(),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=value)

    # Column widths
    for col, width in zip(range(1, len(headers) + 1), [18, 30, 12, 14, 14, 8, 10, 10, 28, 15, 12]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ── Build filename e.g. CSE_2023-2027_4thYear_A.xlsx ──────────────────
    dept_code = (profile.department.code if not profile.is_super_admin
                 else (qs.first().department.code if qs.exists() else 'All'))
    parts = [dept_code]
    if filters.get('batch'):
        parts.append(str(filters['batch']).replace('-', '_'))
    if filters.get('year'):
        parts.append(f"{filters['year']}Year")
    if filters.get('section'):
        parts.append(str(filters['section']))
    filename = f"{'_'.join(parts)}.xlsx"

    AuditLog.log(
        user=request.user,
        action=AuditLog.ACTION_EXCEL_EXPORT,
        description=f"Exported {qs.count()} students to Excel (filters: {filters or 'none'})",
        department=(profile.department if not profile.is_super_admin else None),
        ip_address=request.client_ip,
        user_agent=request.user_agent,
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
